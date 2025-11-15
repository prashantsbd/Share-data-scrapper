import openpyxl
import selenium
import time
from openpyxl.styles import NamedStyle, Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common import keys
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook
from selenium.webdriver.chrome import options as option
from time import sleep


def actualcompany(driver):
    return len(driver.find_element(By.XPATH, "//button[@class='dropdown-item active']").text) == len(stock)


def url_changed(driver):
    return driver.current_url != initial_url


def lpt_not_zero(driver):
    return driver.find_element(By.XPATH, "//tbody/tr[4]/td/div/span[1]").text != '0'


start_time = time.time()

file = "C:/Users/ritik/OneDrive/Desktop/sharescrap.xlsx"
workbook = load_workbook(file)
worksheet = workbook["sharegyan"]


option = webdriver.ChromeOptions()
option.add_argument("--windows-size=1100,660")
option.add_argument("--disable-notification")
option.add_argument("--disable-blink-feature=AutomationControlled")

driver = webdriver.Chrome(options=option)
driver.get("https://nepalstock.com/company")
driver.find_element(By.XPATH, "//div[@class='table__perpage']/select[1]").click()
driver.implicitly_wait(5)
driver.find_element(By.XPATH, "//div[@class='table__perpage']/select/option[6]").click()
calibri_font = Font(name="Calibri")
driver.implicitly_wait(5)
chiz = len(driver.find_elements(By.XPATH, "//tr"))
for i in range(1, chiz):
    symbol = driver.find_element(By.XPATH, f"//tr[{i}]/td[3]").text
    name = driver.find_element(By.XPATH, f"//tr[{i}]/td[2]").text
    stock = f"({symbol}) {name}"
    cell = worksheet.cell(row=i+1, column=1, value=stock)
    cell.font = calibri_font


driver.get("https://nepalstock.com/")
try:
    driver.implicitly_wait(2)
    driver.find_element(By.XPATH, "//button[@class='close']").click()
except:
    pass
pub = 1
i = 2
a = 1
for material in worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
    if material[0] is None:
        break
    stock = worksheet.cell(i, 1).value
    initial_url = driver.current_url
    driver.find_element(By.XPATH, "//input").send_keys(stock)
    try:
        WebDriverWait(driver, 5).until(actualcompany)
    except:
        pass
    try:
        driver.find_element(By.XPATH, "//button[@class='dropdown-item active']").click()
    except:
        stocks = stock.split(" ")
        stock = stocks[0]
        driver.refresh()
        sleep(3)
        driver.find_element(By.XPATH, "//input").send_keys(stock)
        sleep(2)
        driver.find_element(By.XPATH, "//button[@class='dropdown-item active']").click()
    try:
        WebDriverWait(driver, 5).until(url_changed)
    except:
        driver.refresh()
        sleep(2)
        driver.refresh()
        sleep(2)
        driver.find_element(By.XPATH, "//input").send_keys(stock)
        sleep(1.5)
        driver.find_element(By.XPATH, "//button[@class='dropdown-item active']").click()
        WebDriverWait(driver, 5).until(url_changed)
    try:
        WebDriverWait(driver, 2).until(lpt_not_zero)
        while pub == driver.find_element(By.XPATH, "//div[@id='companytabcontent']//tr[3]/td").text:
            sleep(0.45)
        pub = driver.find_element(By.XPATH, "//div[@id='companytabcontent']//tr[3]/td").text
        price = driver.find_element(By.XPATH, "//tbody/tr[4]/td/div/span[1]").text
    except:
        i += 1
        continue
    if price != driver.find_element(By.XPATH, "//tbody/tr[4]/td/div/span[1]").text:
        price = driver.find_element(By.XPATH, "//tbody/tr[4]/td/div/span[1]").text
    ltp = price.replace(",", "")
    public = pub.split(" ")
    worksheet.cell(i, 2, float(ltp))
    worksheet.cell(i, 3, public[0])
    print(i)
    i += 1
workbook.save(file)
workbook.close()
time_taken = (time.time()-start_time)/60
print(f"this task took {time_taken} min")
quit()